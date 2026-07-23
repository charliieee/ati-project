"""Adjudication workbook for the coder consensus meeting.

One row per unresolved case, with the personas to decide between and a
guiding question tailored to each rule combination.
"""

# ===== IMPORTS =====
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import persona_clustering as pc
from assignment_rules import ACTIVE, apply_rules

# ===== CONFIGURATION & CONSTANTS =====
PERSONAS = ["Industry Mover", "Policy Shaper", "Academic Researcher",
            "Public Communicator", "Frontline Clinician"]

QUESTIONS = {  # keyed by the set of matched rules
    ("Industry Mover", "Public Communicator"):
        "Liegt der Schwerpunkt auf dem Aufbau von Unternehmen/Produkten oder auf dem öffentlichen Diskursbeitrag?",
    ("Policy Shaper", "Public Communicator"):
        "Wirkt die Person primär institutionell (Gremien, Advisory) oder über Bühne/Öffentlichkeit?",
    ("Industry Mover", "Academic Researcher"):
        "Ist die AI-Arbeit primär kommerziell (Firma, Produkt) oder wissenschaftlich (Forschung, Fellowship) verankert?",
    ("Academic Researcher", "Public Communicator"):
        "Dominiert Originalforschung oder der Debatten-/Review-Beitrag zum Diskurs?",
    ("Policy Shaper", "Academic Researcher"):
        "Fließt die Forschungsarbeit primär in Gremien/Policy oder bleibt sie akademisch?",
    ("Industry Mover", "Policy Shaper", "Public Communicator"):
        "Portfolio-Profil: Welche der drei Einflusslogiken (Unternehmen / Gremien / Öffentlichkeit) prägt das Gesamtbild?",
    ("Policy Shaper", "Academic Researcher", "Public Communicator"):
        "Welche Logik prägt das Gesamtbild: Gremien, Forschung oder öffentlicher Diskurs?",
    (): "Kein Regel-Treffer: Gesamtbild qualitativ zuordnen (Interview-/Footprint-Kontext heranziehen).",
}

HEAD = ["ID", "Kohorte", "Entscheidung zwischen", "Leitfrage",
        "Aktive Merkmale", "Karriere (Kurzinfo)", "Konsens-Persona", "Begründung"]
WIDTHS = [6, 9, 34, 52, 36, 44, 22, 44]

# ===== CORE FUNCTIONS =====

def build_rows(df, hits, assigned):
    work = pd.read_excel(pc.HERE / "data" / "Overview_evaluation_cleaned_Claude.xlsx",
                         sheet_name="Work (all Cohorts)")
    work.columns = [str(c).strip() for c in work.columns]
    work["ID"] = pd.to_numeric(work["ID"], errors="coerce")
    career = work.set_index("ID")["Career after Fellowship"]
    rows = []
    for idx in df.index[assigned == "ADJUDICATION"]:
        matched = tuple(n for n in hits.columns if hits.loc[idx, n])
        options = " oder ".join(matched) if matched else "alle 5 Personas (offen)"
        rows.append([int(df.ID[idx]), int(df.Cohort[idx]), options,
                     QUESTIONS[matched],
                     ", ".join(c for c in ACTIVE if df.loc[idx, c] == 1),
                     str(career.get(df.ID[idx], ""))[:180], "", ""])
    return rows

# ===== MAIN EXECUTION =====

def main():
    df = pc.load(pc.HERE / "data" / "Overview_evaluation_cleaned_Claude.xlsx")
    hits, assigned = apply_rules(df)
    rows = build_rows(df, hits, assigned)

    wb = Workbook()
    ws = wb.active
    ws.title = "Adjudikation"
    yellow = PatternFill("solid", fgColor="FFFF00")
    grey = Font(name="Arial", italic=True, color="808080", size=9)

    ws["A1"] = ("Konsensrunde Persona-Zuordnung — nur die gelben Spalten ausfüllen: "
                "'Konsens-Persona' (Dropdown) und 'Begründung' (1–2 Sätze, dient als Audit-Trail).")
    ws["A1"].font = Font(name="Arial", bold=True, size=10)
    ws["A2"] = ("Beispiel: ID 99 | Entscheidung zwischen Policy Shaper oder Public Communicator | "
                "Konsens-Persona: Policy Shaper | Begründung: NICE-Committee ist Hauptaktivität, "
                "Vorträge nur begleitend.")
    ws["A2"].font = grey

    ws.append(HEAD)
    for c in ws[3]:
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
    for r in rows:
        ws.append(r)

    body = ws.iter_rows(min_row=4, max_row=3 + len(rows))
    for row in body:
        for c in row:
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        row[6].fill = yellow
        row[7].fill = yellow

    dv = DataValidation(type="list", formula1='"' + ",".join(PERSONAS) + '"',
                        allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(f"G4:G{3 + len(rows)}")

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    out = pc.OUT / "Adjudikation_Konsensrunde.xlsx"
    wb.save(out)
    print(f"{len(rows)} Fälle -> {out}")


if __name__ == "__main__":
    main()
